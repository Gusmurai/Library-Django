from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.utils import timezone        # Для получения текущего времени
from datetime import timedelta
from .models import Book, Booking, Reader, RejectionReason, Publisher, Favorite
from .models import News, Genre, Author,User, LibraryInfo
from django.db.models import Q, ProtectedError
from .forms import UserEditForm, LibraryInfoForm, NewsForm, RejectionReasonForm, AuthorForm, \
    GenreForm, PublisherForm, BookForm, AdminSetPasswordForm  # Импортируем форму
from django.contrib.auth.hashers import make_password
from .forms import UserCreateForm, UserEditForm
from .forms import ProfileUpdateForm, ChangePasswordForm
from django.contrib.auth import update_session_auth_hash
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse, FileResponse
from django.db.models import Count
import io
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4


def catalog_view(request):
    active_tab = request.GET.get('tab', 'books')

    all_genres = Genre.objects.all().order_by('name')
    all_authors = Author.objects.all().order_by('last_name')
    all_publishers = Publisher.objects.all().order_by('name')

    # Получение поискового запроса для справочников
    q_dict = request.GET.get('q_dict', '')
    # Динамическая фильтрация справочников при вводе данных в поиск
    if q_dict:
        if active_tab == 'authors':
            all_authors = all_authors.filter(
                Q(last_name__icontains=q_dict) |
                Q(first_name__icontains=q_dict) |
                Q(middle_name__icontains=q_dict)
            )
        elif active_tab == 'genres':
            all_genres = all_genres.filter(name__icontains=q_dict)
        elif active_tab == 'publishers':
            all_publishers = all_publishers.filter(name__icontains=q_dict)

    # Инициализация форм справочников
    author_form = AuthorForm()
    genre_form = GenreForm()
    publisher_form = PublisherForm()

    # Обработка ВВОДА данных (ТОЛЬКО ДЛЯ БИБЛИОТЕКАРЯ)
    if request.method == 'POST' and request.user.is_authenticated and request.user.role == 'librarian':
        if 'add_author' in request.POST:
            author_form = AuthorForm(request.POST)
            if author_form.is_valid():
                author_form.save()
                messages.success(request, "Автор успешно добавлена")
                return redirect('/catalog/?tab=authors')
            else:
                active_tab = 'authors'  # Оставляем вкладку открытой, чтобы показать ошибки

        elif 'add_genre' in request.POST:
            genre_form = GenreForm(request.POST)
            if genre_form.is_valid():
                genre_form.save()
                messages.success(request, "Жанр успешно добавлен")
                return redirect('/catalog/?tab=genres')
            else:
                active_tab = 'genres'

        elif 'add_publisher' in request.POST:
            publisher_form = PublisherForm(request.POST)
            if publisher_form.is_valid():
                publisher_form.save()
                messages.success(request, "Издательство успешно добавлено")
                return redirect('/catalog/?tab=publishers')
            else:
                active_tab = 'publishers'

    # Логика вывода книг (Библиотекарь видит архив во вкладке)
    if active_tab == 'archive' and request.user.role in ['librarian', 'admin']:
        books = Book.objects.filter(is_archived=True)
    else:
        books = Book.objects.filter(is_archived=False)

    # Фильтрация
    query = request.GET.get('q')
    genre_id = request.GET.get('genre')
    issue_type = request.GET.get('issue_type')

    if query:
        books = books.filter(Q(title__icontains=query) | Q(authors__last_name__icontains=query) | Q(
            authors__first_name__icontains=query))
    if genre_id:
        books = books.filter(genres__id=genre_id)
    if issue_type:
        books = books.filter(issue_type=issue_type)

    return render(request, 'library/catalog.html', {
        'books': books.distinct().order_by('title'),
        'all_genres': all_genres,
        'all_authors': all_authors,
        'all_publishers': all_publishers,
        'active_tab': active_tab,
        'q_dict': q_dict,
        # Передаем формы в шаблон
        'author_form': author_form,
        'genre_form': genre_form,
        'publisher_form': publisher_form,
    })


@login_required
def cancel_booking_view(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    # Проверка: отменить может только владелец и только активную бронь
    if booking.reader.user == request.user and booking.status in ['pending', 'approved']:
        booking.status = 'cancelled'
        booking.status_change_date = timezone.now()
        booking.save()
        messages.success(request, f"Вы успешно отменили бронирование книги «{booking.book.title}»")
    else:
        messages.error(request, "Эту бронь нельзя отменить.")

    return redirect('/profile/?tab=dashboard')
# library/views.py

# @login_required
def catalog_view(request):
    # 1. СРАЗУ инициализируем все переменные, которые пойдут в шаблон.
    # Это гарантирует, что мы не получим ошибку UnboundLocalError.
    active_tab = request.GET.get('tab', 'books')
    user_favorite_ids = []
    q_dict = request.GET.get('q_dict', '')

    # Инициализируем пустые формы
    author_form = AuthorForm()
    genre_form = GenreForm()
    publisher_form = PublisherForm()

    # 2. ПОЛУЧАЕМ СПРАВОЧНИКИ (базовые списки)
    all_genres = Genre.objects.all().order_by('name')
    all_authors = Author.objects.all().order_by('last_name')
    all_publishers = Publisher.objects.all().order_by('name')

    # 3. ЛОГИКА ИЗБРАННОГО (получаем ID книг для сердечек)
    if request.user.is_authenticated and request.user.role == 'reader':
        if hasattr(request.user, 'reader_profile'):
            user_favorite_ids = request.user.reader_profile.favorites.values_list('book_id', flat=True)

    # 4. ОБРАБОТКА ПОИСКА ПО СПРАВОЧНИКАМ (Авторы, Жанры, Издательства)
    if q_dict:
        if active_tab == 'authors':
            all_authors = all_authors.filter(
                Q(last_name__icontains=q_dict) |
                Q(first_name__icontains=q_dict) |
                Q(middle_name__icontains=q_dict)
            )
        elif active_tab == 'genres':
            all_genres = all_genres.filter(name__icontains=q_dict)
        elif active_tab == 'publishers':
            all_publishers = all_publishers.filter(name__icontains=q_dict)

    # 5. ОБРАБОТКА POST-ЗАПРОСОВ (Добавление библиотекарем)
    if request.method == 'POST' and request.user.role == 'librarian':
        if 'add_author' in request.POST:
            author_form = AuthorForm(request.POST)
            if author_form.is_valid():
                author_form.save()
                messages.success(request, "Автор успешно добавлен")
                return redirect('/catalog/?tab=authors')
            active_tab = 'authors' # Если ошибка, остаемся на вкладке

        elif 'add_genre' in request.POST:
            genre_form = GenreForm(request.POST)
            if genre_form.is_valid():
                genre_form.save()
                messages.success(request, "Жанр успешно добавлен")
                return redirect('/catalog/?tab=genres')
            active_tab = 'genres'

        elif 'add_publisher' in request.POST:
            publisher_form = PublisherForm(request.POST)
            if publisher_form.is_valid():
                publisher_form.save()
                messages.success(request, "Издательство успешно добавлено")
                return redirect('/catalog/?tab=publishers')
            active_tab = 'publishers'

    # 6. ЛОГИКА ВЫБОРА КНИГ (в зависимости от вкладки)
    if active_tab == 'favorites' and request.user.role == 'reader':
        if hasattr(request.user, 'reader_profile'):
            # Фильтруем через связь с моделью Favorite (связь с Reader)
            books = Book.objects.filter(favored_by__reader=request.user.reader_profile, is_archived=False)
        else:
            books = Book.objects.none()
    elif active_tab == 'archive' and request.user.role in ['librarian', 'admin']:
        books = Book.objects.filter(is_archived=True)
    else:
        # Обычный каталог
        books = Book.objects.filter(is_archived=False)

    # 7. ФИЛЬТРАЦИЯ И ПОИСК ПО КНИГАМ (общий поиск)
    query = request.GET.get('q')
    genre_id = request.GET.get('genre')
    issue_type = request.GET.get('issue_type')

    if query:
        books = books.filter(
            Q(title__icontains=query) |
            Q(authors__last_name__icontains=query) |
            Q(authors__first_name__icontains=query)
        )
    if genre_id:
        books = books.filter(genres__id=genre_id)
    if issue_type:
        books = books.filter(issue_type=issue_type)

    # 8. ФИНАЛЬНЫЙ РЕНДЕРИНГ
    return render(request, 'library/catalog.html', {
        'books': books.distinct().order_by('title'),
        'all_genres': all_genres,
        'all_authors': all_authors,
        'all_publishers': all_publishers,
        'active_tab': active_tab,
        'q_dict': q_dict,
        'author_form': author_form,
        'genre_form': genre_form,
        'publisher_form': publisher_form,
        'user_favorite_ids': user_favorite_ids,
    })


@login_required
def toggle_favorite_view(request, book_id):
    # 1. Проверяем, что действие совершает именно читатель
    if request.user.role != 'reader':
        messages.error(request, "Только читатели могут добавлять книги в избранное.")
        return redirect('library:catalog')

    # 2. Получаем книгу
    book = get_object_or_404(Book, id=book_id)

    # 3. Получаем профиль читателя (через связанную модель ReaderProfile)
    reader_profile = getattr(request.user, 'reader_profile', None)
    if not reader_profile:
        messages.error(request, "Профиль читателя не найден.")
        return redirect('library:catalog')

    # 4. Логика переключения (Toggle)
    # get_or_create возвращает кортеж: (объект, created_boolean)
    fav, created = Favorite.objects.get_or_create(reader=reader_profile, book=book)

    if not created:
        # Если запись уже существовала, удаляем её (убираем из избранного)
        fav.delete()
        messages.info(request, f"Книга «{book.title}» удалена из избранного.")
    else:
        # Если запись создалась только что — значит добавили
        messages.success(request, f"Книга «{book.title}» добавлена в избранное!")

    # 5. Возвращаемся на ту страницу, откуда пришел пользователь
    return redirect(request.META.get('HTTP_REFERER', 'library:catalog'))


def book_detail_view(request, book_id):
    book = get_object_or_404(Book, pk=book_id)

    # Инициализируем переменную заранее
    user_favorite_ids = []

    if request.user.is_authenticated and request.user.role == 'reader':
        if hasattr(request.user, 'reader_profile'):
            user_favorite_ids = request.user.reader_profile.favorites.values_list('book_id', flat=True)

    return render(request, 'library/book_detail.html', {
        'book': book,
        'user_favorite_ids': user_favorite_ids
    })

# Функция динамического ввода новой заявки на бронирование
@login_required
def book_booking_view(request, book_id):
    if request.method == 'POST':
        book = get_object_or_404(Book, pk=book_id)

        # Получаем профиль читателя текущего пользователя
        # Если пользователь - не читатель (например, админ), бронирование "запрещено"
        try:
            reader = request.user.reader_profile
        except:
            messages.error(request, "Бронирование доступно только читателям")
            return redirect('library:catalog')

        # Проверка бизнес-правила: не более 5 активных броней
        active_bookings = Booking.objects.filter(
            reader=reader,
            status__in=['pending', 'approved']
        ).count()

        if active_bookings < 5:
            # Динамический ввод записи в базу данных
            Booking.objects.create(reader=reader, book=book, status='pending')
            messages.success(request, f"Заявка на книгу '{book.title}' успешно создана!")
        else:
            messages.error(request, "Достигнут лимит: нельзя иметь более 5 активных заявок")

    return redirect('library:catalog')


# Функция динамического вывода ленты новостей с поддержкой поиска
def news_view(request):
    query = request.GET.get('q')
    if query:
        news_list = News.objects.filter(
            Q(title__icontains=query) |
            Q(short_description__icontains=query) |
            Q(full_description__icontains=query)
        ).distinct().order_by('-publish_date')
    else:
        news_list = News.objects.all().order_by('-publish_date')

    # Логика добавления новости (только для персонала)
    form = None
    if request.user.is_authenticated and request.user.role in['admin', 'librarian']:
        if request.method == 'POST' and 'add_news' in request.POST:
            form = NewsForm(request.POST, request.FILES)
            if form.is_valid():
                new_news = form.save(commit=False)
                new_news.author = request.user # Автором становится тот, кто нажал кнопку
                new_news.save()
                messages.success(request, "Новость успешно опубликована!")
                return redirect('library:news')
        else:
            form = NewsForm()

    return render(request, 'library/news.html', {'news': news_list, 'form': form})

# Страница "О нас" с динамической статистикой
def about_view(request):
    # Получаем единственную запись с настройками
    info = LibraryInfo.objects.first()

    # Считаем данные для блока статистики
    stats = {
        'total_books': Book.objects.filter(is_archived=False).count(),
        'active_bookings': Booking.objects.filter(status__in=['pending', 'approved']).count(),
        'completed_bookings': Booking.objects.filter(status='completed').count(),
    }

    return render(request, 'library/about.html', {
        'info': info,
        'stats': stats
    })
# Функция динамического вывода личного кабинета пользователя
@login_required
def profile_view(request):
    user = request.user
    # Получаем вкладку из URL, по умолчанию 'dashboard'
    active_tab = request.GET.get('tab', 'dashboard')
    status_filter = request.GET.get('filter', 'all')  # Получаем фильтр

    profile_form = ProfileUpdateForm(instance=user)
    password_form = ChangePasswordForm()

    if 'update_profile' in request.POST:
        profile_form = ProfileUpdateForm(request.POST, instance=user)
        if profile_form.is_valid():
            # Сохраняем основные данные пользователя
            user_obj = profile_form.save(commit=False)
            # Принудительно возвращаем старый логин (защита от изменения PK)
            user_obj.username = request.user.username
            user_obj.save()

            # Если это читатель — сохраняем подписку
            if user_obj.role == 'reader':
                # Получаем значение чекбокса ('on' — если стоит галочка)
                subscribed = request.POST.get('is_subscribed') == 'on'

                # Обновляем или создаем профиль читателя
                Reader.objects.update_or_create(
                    user=user_obj,
                    defaults={'is_subscribed': subscribed}
                )

            messages.success(request, "Контактные данные успешно обновлены")
            return redirect('/profile/?tab=settings')


        elif 'change_password' in request.POST:

            password_form = ChangePasswordForm(request.POST)

            if password_form.is_valid():

                # Проверяем, ввел ли пользователь данные (если поля пустые, clean вернул None/пустые строки)

                old_pass = password_form.cleaned_data.get('old_password')

                if old_pass:  # Пароль меняется только если заполнено поле старого пароля

                    if user.check_password(old_pass):

                        user.set_password(password_form.cleaned_data['new_password'])

                        user.save()

                        update_session_auth_hash(request, user)

                        messages.success(request, "Пароль успешно изменен")

                        return redirect('/profile/?tab=settings')

                    else:

                        messages.error(request, "Неверный текущий пароль")

                else:

                    # Пользователь ничего не ввел, просто игнорируем

                    pass
            # Если форма невалидна, тоже остаемся на вкладке настроек
            active_tab = 'settings'

    # Сбор статистики
    context = {'active_tab': active_tab, 'filter_val': status_filter}

    if user.role == 'admin':
        context['staff_stats'] = {
            'total_users': User.objects.count(),
            'total_books': Book.objects.count(),
            'pending_bookings': Booking.objects.filter(status='pending').count(),
            'approved_bookings': Booking.objects.filter(status='approved').count(),
            'completed_bookings': Booking.objects.filter(status='completed').count(),
            'cancelled_by_reader': Booking.objects.filter(status='cancelled').count(),
            'rejected_by_librarian': Booking.objects.filter(status='rejected').count(),
            'expired_bookings': Booking.objects.filter(status='expired').count(),
            'all_time_bookings': Booking.objects.count(),
        }
    elif user.role == 'librarian':
        # НОВАЯ СТАТИСТИКА БИБЛИОТЕКАРЯ
        context['staff_stats'] = {
            'approved_by_me': Booking.objects.filter(librarian=user, status='approved').count(),
            'completed_by_me': Booking.objects.filter(librarian=user, status='completed').count(),
            'cancelled_by_me': Booking.objects.filter(librarian=user, status__in=['rejected', 'expired']).count(),
        }
    else:
        # Читатель
        try:
            reader_profile = user.reader_profile
            bookings_query = Booking.objects.filter(reader=reader_profile).select_related('book',
                                                                                          'rejection_type').order_by(
                '-booking_date')

            # Рабочая фильтрация
            if status_filter == 'active':
                bookings_query = bookings_query.filter(status__in=['pending', 'approved'])
            elif status_filter != 'all':
                bookings_query = bookings_query.filter(status=status_filter)

            reader_stats = {
                'active': Booking.objects.filter(reader=reader_profile, status__in=['pending', 'approved']).count(),
                'completed': Booking.objects.filter(reader=reader_profile, status='completed').count(),
                'total': Booking.objects.filter(reader=reader_profile).count(),
            }
        except:  # Если это админ/либ без профиля читателя
            bookings_query = []
            reader_stats = {}

        context['bookings'] = bookings_query
        context['reader_stats'] = reader_stats

    context.update({
        'profile_form': profile_form,
        'password_form': password_form,
    })
    return render(request, 'library/profile.html', context)
def contacts_view(request):
    info = LibraryInfo.objects.first()
    return render(request, 'library/contacts.html', {'info': info})

# Функция динамического вывода полной статьи новости
def news_detail_view(request, news_id):
    # Получаем конкретную новость или выдаем 404, если её нет
    item = get_object_or_404(News, pk=news_id)
    return render(request, 'library/news_detail.html', {'item': item})


@login_required
def admin_panel_view(request):
    """
    Управление пользователями (Админ) и Читателями (Библиотекарь).
    Реализует поиск, фильтрацию и создание новых записей.
    """
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    # Инициализация формы добавления (пустая по умолчанию)
    form = UserCreateForm()

    # 1. ЛОГИКА ДОБАВЛЕНИЯ ПОЛЬЗОВАТЕЛЯ
    if request.method == 'POST' and 'add_user' in request.POST:
        form = UserCreateForm(request.POST)
        if form.is_valid():
            new_user = form.save(commit=False)
            new_user.set_password(form.cleaned_data['password'])

            # Ограничение роли согласно ТЗ
            if request.user.role == 'librarian':
                new_user.role = 'reader'

            new_user.save()

            # Если создан читатель — создаем ему запись в таблице Reader
            if new_user.role == 'reader':
                ticket = form.cleaned_data.get('ticket_number')
                subscribed = form.cleaned_data.get('is_subscribed')

                Reader.objects.create(
                    user=new_user,
                    ticket_number=ticket if ticket else f"ID-{new_user.username}",
                    is_subscribed=subscribed
                )

            messages.success(request, f"Пользователь {new_user.username} успешно создан")
            return redirect('library:admin_panel')
        else:
            # Форма невалидна — сообщения об ошибках отобразятся в шаблоне
            messages.error(request, "Ошибка при создании. Проверьте правильность введенных данных.")

    # 2. ФОРМИРОВАНИЕ СПИСКА (РАЗГРАНИЧЕНИЕ ДОСТУПА)
    if request.user.role == 'admin':
        # Админ управляет только персоналом (кроме себя)
        users_query = User.objects.filter(role__in=['admin', 'librarian']).exclude(username=request.user.username)
    else:
        # Библиотекарь управляет только читателями
        users_query = User.objects.filter(role='reader')

    # 3. ПОИСК И ФИЛЬТРАЦИЯ
    search_query = request.GET.get('q')
    status_filter = request.GET.get('status')
    role_filter = request.GET.get('role_filter')

    if search_query:
        users_query = users_query.filter(
            Q(last_name__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(reader_profile__ticket_number__icontains=search_query)  # Поиск по билету через связь
        )

    if status_filter == 'active':
        users_query = users_query.filter(is_active=True)
    elif status_filter == 'blocked':
        users_query = users_query.filter(is_active=False)

    if request.user.role == 'admin' and role_filter and role_filter != 'all':
        users_query = users_query.filter(role=role_filter)

    return render(request, 'library/admin_panel.html', {
        'users_list': users_query.distinct().order_by('last_name'),
        'form': form,
        'is_admin': request.user.role == 'admin'
    })

@login_required
def toggle_user_status_view(request, username):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    target_user = get_object_or_404(User, username=username)
    target_user.is_active = not target_user.is_active
    target_user.save()

    status_text = "разблокирован" if target_user.is_active else "заблокирован"
    messages.success(request, f"Пользователь {username} {status_text}")
    return redirect('library:admin_panel')


@login_required
def librarian_view(request):
    if request.user.role not in ['librarian', 'admin']:
        return redirect('library:catalog')

    active_tab = request.GET.get('tab', 'active')

    # 1. Автоматическая проверка просрочки
    expired_limit = timezone.now() - timedelta(days=3)
    Booking.objects.filter(status='approved', status_change_date__lt=expired_limit).update(status='expired')

    # 2. Базовый запрос
    all_bookings_query = Booking.objects.all().select_related('reader__user', 'book', 'librarian', 'rejection_type')

    # 3. Фильтры
    query = request.GET.get('q')
    status_filter = request.GET.get('status')
    staff_filter = request.GET.get('staff')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if query:
        all_bookings_query = all_bookings_query.filter(
            Q(book__title__icontains=query) |
            Q(reader__user__last_name__icontains=query) |
            Q(reader__ticket_number__icontains=query)
        )
    if status_filter and status_filter != 'all':
        all_bookings_query = all_bookings_query.filter(status=status_filter)
    if staff_filter and staff_filter != 'all':
        all_bookings_query = all_bookings_query.filter(librarian__username=staff_filter)

    # Фильтр по периоду дат
    if start_date:
        all_bookings_query = all_bookings_query.filter(status_change_date__date__gte=start_date)
    if end_date:
        all_bookings_query = all_bookings_query.filter(status_change_date__date__lte=end_date)

    # 4. Разделение списков
    active_bookings = all_bookings_query.filter(status__in=['pending', 'approved']).order_by('-booking_date')
    history_bookings = all_bookings_query.exclude(status__in=['pending', 'approved']).order_by('-status_change_date')

    # ВАЖНО: Фильтруем ТОЛЬКО библиотекарей для выпадающего списка
    staff_list = User.objects.filter(role='librarian')

    reasons = RejectionReason.objects.all()
    reason_form = RejectionReasonForm()

    return render(request, 'library/librarian.html', {
        'active_bookings': active_bookings,
        'history_bookings': history_bookings,
        'reasons': reasons,
        'reason_form': reason_form,
        'staff_list': staff_list,
        'active_tab': active_tab,
        'q': query,
        'status_val': status_filter,
        'staff_val': staff_filter
    })

# Исправленная функция удаления
@login_required
def delete_dict_item(request, item_type, item_id):
    if request.method != 'POST' or request.user.role != 'librarian':
        return redirect('library:catalog')

    # Инициализация переменных по умолчанию
    tab = 'books'
    base_url = '/catalog/'

    try:
        if item_type == 'author':
            obj = get_object_or_404(Author, id=item_id)
            if obj.book_set.exists():
                messages.error(request, f"Нельзя удалить автора «{obj}», т.к. за ним числятся книги.")
            else:
                obj.delete()
                messages.success(request, "Автор успешно удален.")
            tab = 'authors'

        elif item_type == 'genre':
            obj = get_object_or_404(Genre, id=item_id)
            if obj.book_set.exists():
                messages.error(request, f"Нельзя удалить жанр «{obj.name}», т.к. в нем есть книги.")
            else:
                obj.delete()
                messages.success(request, "Жанр успешно удален.")
            tab = 'genres'

        elif item_type == 'publisher':
            obj = get_object_or_404(Publisher, id=item_id)
            # Здесь сработает автоматический PROTECT, если есть книги
            obj.delete()
            messages.success(request, "Издательство успешно удалено.")
            tab = 'publishers'

        elif item_type == 'reason': # ДОБАВИЛИ ПОДДЕРЖКУ ПРИЧИН
            obj = get_object_or_404(RejectionReason, id=item_id)
            obj.delete()
            messages.success(request, "Причина отказа удалена.")
            tab = 'reasons'
            base_url = '/librarian/'

    except ProtectedError:
        messages.error(request, "Удаление невозможно: эта запись используется в каталоге.")
    except Exception as e:
        messages.error(request, f"Ошибка при удалении: {e}")

    return redirect(f"{base_url}?tab={tab}")


@login_required
def edit_user_view(request, username):
    # 1. Проверка прав
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    # 2. Получение объекта пользователя
    edit_user = get_object_or_404(User, username=username)
    reader_profile = getattr(edit_user, 'reader_profile', None)

    # 3. Инициализация форм по умолчанию (для GET и как запасной вариант для POST)
    initial_data = {}
    if reader_profile:
        initial_data['ticket_number'] = reader_profile.ticket_number
        initial_data['is_subscribed'] = reader_profile.is_subscribed

    form = UserEditForm(instance=edit_user, initial=initial_data)
    password_form = AdminSetPasswordForm()

    # 4. Обработка POST-запроса
    if request.method == 'POST':
        # --- Смена пароля ---
        if 'change_password' in request.POST:
            password_form = AdminSetPasswordForm(request.POST)
            if password_form.is_valid():
                edit_user.set_password(password_form.cleaned_data['new_password'])
                edit_user.save()
                messages.success(request, f"Пароль для {edit_user.username} успешно изменен.")
                return redirect('library:edit_user', username=username)
            else:
                messages.error(request, "Ошибка смены пароля: проверьте, совпадают ли пароли.")

        # --- Редактирование профиля ---
        elif 'update_profile' in request.POST:
            form = UserEditForm(request.POST, instance=edit_user)
            if form.is_valid():
                user_obj = form.save(commit=False)
                user_obj.username = username  # Фиксируем логин
                user_obj.save()

                if user_obj.role == 'reader':
                    ticket = form.cleaned_data.get('ticket_number')
                    subscribed = form.cleaned_data.get('is_subscribed')
                    Reader.objects.update_or_create(
                        user=user_obj,
                        defaults={
                            'ticket_number': ticket,
                            'is_subscribed': subscribed
                        }
                    )
                messages.success(request, "Данные пользователя успешно обновлены.")
                return redirect('library:admin_panel')
            else:
                messages.error(request, "Ошибка сохранения профиля. Проверьте введенные данные.")

    # 5. Рендеринг шаблона — теперь обе формы точно определены
    return render(request, 'library/edit_user.html', {
        'form': form,
        'password_form': password_form,
        'edit_user': edit_user
    })
@login_required
def news_edit_view(request, news_id=None):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:news')

    news_item = get_object_or_404(News, pk=news_id) if news_id else None

    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES, instance=news_item)
        if form.is_valid():
            new_news = form.save(commit=False)
            if not news_id:
                new_news.author = request.user
            new_news.save()
            messages.success(request, "Новость успешно сохранена")
            return redirect('library:news')
    else:
        form = NewsForm(instance=news_item)

    return render(request, 'library/news_form.html', {'form': form, 'edit': news_id})


@login_required
def news_delete_view(request, news_id):
    news_item = get_object_or_404(News, pk=news_id)
    # Правило: админ удаляет всё, библиотекарь только своё
    if request.user.role == 'admin' or news_item.author == request.user:
        news_item.delete()
        messages.success(request, "Новость удалена")
    else:
        messages.error(request, "Вы можете удалять только свои новости")
    return redirect('library:news')


# --- Настройки сайта (только Админ) ---
@login_required
def settings_view(request):
    if request.user.role != 'admin':
        messages.error(request, "Доступ запрещен. Требуются права администратора.")
        return redirect('library:catalog')

        # Берем первую запись настроек или создаем пустой объект
    info = LibraryInfo.objects.first()

    if request.method == 'POST':
        # Если info существует, мы его обновляем (instance=info), если нет - создаем новое
        form = LibraryInfoForm(request.POST, instance=info)
        if form.is_valid():
            new_info = form.save(commit=False)
            new_info.admin = request.user  # Фиксируем, кто последним менял настройки
            new_info.save()
            messages.success(request, "Настройки библиотеки успешно обновлены!")
            return redirect('library:settings')
    else:
        # Загружаем текущие данные в форму
        form = LibraryInfoForm(instance=info)

    return render(request, 'library/settings.html', {'form': form, 'info': info})

@login_required
def news_edit_view(request, news_id):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:news')

    news_item = get_object_or_404(News, id=news_id)

    # Запрещаем библиотекарю редактировать чужие новости
    if request.user.role == 'librarian' and news_item.author != request.user:
        messages.error(request, "Вы можете редактировать только свои новости.")
        return redirect('library:news')

    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES, instance=news_item)
        if form.is_valid():
            form.save()
            messages.success(request, "Новость успешно обновлена!")
            return redirect('library:news')
    else:
        form = NewsForm(instance=news_item)

    return render(request, 'library/news_edit.html', {'form': form, 'news_item': news_item})


# 3. Удаление новости
@login_required
def news_delete_view(request, news_id):
    if request.method == 'POST':
        news_item = get_object_or_404(News, id=news_id)

        # Админ удаляет любые, библиотекарь - только свои
        if request.user.role == 'admin' or news_item.author == request.user:
            news_item.delete()
            messages.success(request, "Новость удалена!")
        else:
            messages.error(request, "У вас нет прав на удаление чужой новости.")

    return redirect('library:news')

# Вспомогательная функция для получения полного ФИО
def get_full_fio(user_obj):
    if not user_obj: return "—"
    return f"{user_obj.last_name} {user_obj.first_name} {user_obj.middle_name or ''}".strip()

@login_required
def reports_list_view(request):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    # Для фильтрации по пользователю библиотекарю нужен список читателей
    readers = Reader.objects.all()
    return render(request, 'library/reports.html', {'readers': readers})


@login_required
def report_system_stats_excel(request):
    if request.user.role != 'admin': return redirect('library:catalog')
    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Статистика"
    ws.append(['Показатель', 'Значение'])
    ws.append(['Всего пользователей', User.objects.count()])
    ws.append(['Книг в каталоге', Book.objects.count()])
    ws.append(['Активных заявок', Booking.objects.filter(status__in=['pending', 'approved']).count()])
    ws.append(['Завершено заявок', Booking.objects.filter(status='completed').count()])
    ws.append(['Отменено/Истекло', Booking.objects.filter(status__in=['rejected', 'cancelled', 'expired']).count()])
    for cell in ws[1]: cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=system_stats.xlsx'
    wb.save(response)
    return response


@login_required
def report_rejections_excel(request):
    if request.user.role != 'admin': return redirect('library:catalog')
    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Причины отказов"

    ws.append(
        ['Дата решения', 'Читатель (Билет)', 'Книга', 'Причина', 'Доп. комментарий', 'Библиотекарь (ФИО)'])

    rejections = Booking.objects.filter(status='rejected').select_related('rejection_type', 'reader__user', 'librarian',
                                                                          'book')

    for b in rejections:
        ws.append([
            b.status_change_date.strftime('%d.%m.%Y %H:%M') if b.status_change_date else "—",
            f"{get_full_fio(b.reader.user)} (№{b.reader.ticket_number})",
            b.book.title,
            b.rejection_type.name if b.rejection_type else "Не указана",  # Берем имя из справочника
            b.reject_comment or "—",
            get_full_fio(b.librarian)
        ])

    for cell in ws[1]: cell.font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rejections_analysis.xlsx'
    wb.save(response)
    return response


@login_required
def report_staff_activity_excel(request):
    if request.user.role != 'admin': return redirect('library:catalog')
    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Активность персонала"
    ws.append(['Сотрудник', 'Логин', 'Всего обработано', 'Одобрено', 'Завершено', 'Отклонено'])

    staff = User.objects.filter(role='librarian')
    for s in staff:
        total = Booking.objects.filter(librarian=s).count()
        approved = Booking.objects.filter(librarian=s, status='approved').count()
        completed = Booking.objects.filter(librarian=s, status='completed').count()
        rejected = Booking.objects.filter(librarian=s, status='rejected').count()
        ws.append([get_full_fio(s), s.username, total, approved, completed, rejected])
    for cell in ws[1]: cell.font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=staff_activity.xlsx'
    wb.save(response)
    return response


@login_required
def report_top_books_excel(request):
    if request.user.role != 'admin': return redirect('library:catalog')
    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Рейтинг книг"
    ws.append(['Название книги', 'Авторы', 'Жанры', 'Количество заявок'])

    books = Book.objects.annotate(bookings_count=Count('booking')).order_by('-bookings_count')
    for book in books:
        authors = ", ".join([f"{a.last_name} {a.first_name} {a.middle_name or ''}".strip() for a in book.authors.all()])
        genres = ", ".join([g.name for g in book.genres.all()])
        ws.append([book.title, authors, genres, book.bookings_count])
    for cell in ws[1]: cell.font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=top_books.xlsx'
    wb.save(response)
    return response


# ==========================================
# ОТЧЕТЫ БИБЛИОТЕКАРЯ
# ==========================================

@login_required
def report_booked_period_excel(request):
    if request.user.role != 'librarian': return redirect('library:catalog')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Забронированные за период"
    ws.append(['Дата брони', 'Книга', 'Читатель', 'Номер билета', 'Статус'])

    bookings = Booking.objects.all()
    if start_date: bookings = bookings.filter(booking_date__date__gte=parse_date(start_date))
    if end_date: bookings = bookings.filter(booking_date__date__lte=parse_date(end_date))

    for b in bookings.order_by('-booking_date'):
        ws.append([
            b.booking_date.strftime('%d.%m.%Y'),
            b.book.title,
            get_full_fio(b.reader.user),
            b.reader.ticket_number,
            b.get_status_display()
        ])
    for cell in ws[1]: cell.font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=booked_report.xlsx'
    wb.save(response)
    return response


@login_required
def report_available_books_excel(request):
    if request.user.role != 'librarian': return redirect('library:catalog')
    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = "Полный перечень фонда"

    # ВЫВОДИМ ВСЕ ПОЛЯ КНИГИ (кроме статуса архива, т.к. фильтруем по нему)
    ws.append(['Название', 'Авторы', 'Жанры', 'Издательство', 'Год издания', 'ISBN', 'Тип выдачи', 'Аннотация'])

    books = Book.objects.filter(is_archived=False).order_by('title')
    for b in books:
        authors = ", ".join([f"{a.last_name} {a.first_name} {a.middle_name or ''}".strip() for a in b.authors.all()])
        genres = ", ".join([g.name for g in b.genres.all()])
        ws.append([
            b.title,
            authors,
            genres,
            b.publisher.name if b.publisher else "—",
            b.publish_year or "—",
            b.isbn or "—",
            b.get_issue_type_display(),
            b.description or "—"
        ])

    # Настройка ширины колонок для читаемости
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['H'].width = 50

    for cell in ws[1]: cell.font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=library_catalog_full.xlsx'
    wb.save(response)
    return response


@login_required
def report_user_history_excel(request):
    if request.user.role != 'librarian': return redirect('library:catalog')
    ticket = request.GET.get('ticket')
    if not ticket: return redirect('library:reports_list')

    reader = get_object_or_404(Reader, ticket_number=ticket)
    user_bookings = Booking.objects.filter(reader=reader).order_by('-booking_date')

    wb = openpyxl.Workbook();
    ws = wb.active;
    ws.title = f"Выписка {reader.ticket_number}"

    # 1. ЗАГОЛОВКИ (Твой порядок)
    ws.append(['Дата', 'Книга', 'Статус', 'Причина отказа', 'Комментарий', 'Библиотекарь'])

    # 2. ДАННЫЕ
    for b in user_bookings:
        # Собираем данные
        status_text = b.get_status_display()
        reason = b.rejection_type.name if b.rejection_type else "—"
        comment = b.reject_comment if b.reject_comment else "—"
        librarian = get_full_fio(b.librarian)

        ws.append([
            b.booking_date.strftime('%d.%m.%Y'),
            b.book.title,
            status_text,
            reason,
            comment,
            librarian
        ])

    # Стилизация заголовков
    for cell in ws[1]: cell.font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=user_{ticket}_history.xlsx'
    wb.save(response)
    return response

@login_required
def process_booking_view(request, booking_id, action):
    booking = get_object_or_404(Booking, id=booking_id)

    # Только библиотекарь имеет право менять статус
    if request.user.role != 'librarian':
        return redirect('library:catalog')

    booking.librarian = request.user
    booking.status_change_date = timezone.now()  # Фиксируем время решения

    if action == 'approve':
        booking.status = 'approved'
        # Сохраняем текстовый комментарий (например, про штрафы)
        booking.reject_comment = request.POST.get('staff_comment', '')


    elif action == 'reject':

        reason_id = request.POST.get('reason')

        # ПРОВЕРКА: если причина не выбрана — блокируем действие

        if not reason_id:
            messages.error(request, "Ошибка: выберите причину отказа!")

            return redirect('library:librarian')

        booking.rejection_type = get_object_or_404(RejectionReason, id=reason_id)

        booking.status = 'rejected'

        booking.librarian = request.user

        booking.reject_comment = request.POST.get('staff_comment', '')

    elif action == 'complete':
        booking.status = 'completed'

    booking.save()
    messages.success(request, f"Статус заявки на книгу «{booking.book.title}» изменен.")
    return redirect('library:librarian')


@login_required
def toggle_archive_view(request, book_id):
    if request.user.role != 'librarian':
        return redirect('library:catalog')

    book = get_object_or_404(Book, id=book_id)
    book.is_archived = not book.is_archived  # Переключаем статус
    book.save()

    status_msg = "отправлена в архив" if book.is_archived else "возвращена в фонд"
    messages.success(request, f"Книга «{book.title}» {status_msg}.")

    # Возвращаемся на ту вкладку, где были
    tab = 'archive' if not book.is_archived else 'books'
    return redirect(f'/catalog/?tab={tab}')


# library/views.py

@login_required
def delete_dict_item(request, item_type, item_id):
    # 1. Безопасность: только POST и только Библиотекарь
    if request.method != 'POST' or request.user.role != 'librarian':
        return redirect('library:catalog')

    # 2. СРАЗУ задаем значения по умолчанию (это лечит ошибку UnboundLocalError)
    # Если что-то пойдет не так, мы вернемся просто в каталог
    tab = 'books'
    base_url = '/catalog/'

    # 3. Определяем, куда возвращаться в зависимости от типа
    if item_type == 'reason':
        tab = 'reasons'
        base_url = '/librarian/'
    elif item_type == 'author':
        tab = 'authors'
    elif item_type == 'genre':
        tab = 'genres'
    elif item_type == 'publisher':
        tab = 'publishers'

    # 4. Основная логика удаления
    try:
        if item_type == 'author':
            obj = get_object_or_404(Author, id=item_id)
            if obj.book_set.exists():
                messages.error(request, f"Ошибка: нельзя удалить автора «{obj}», т.к. за ним числятся книги.")
            else:
                obj.delete()
                messages.success(request, "Автор удален.")

        elif item_type == 'genre':
            obj = get_object_or_404(Genre, id=item_id)
            if obj.book_set.exists():
                messages.error(request, f"Ошибка: нельзя удалить жанр «{obj.name}», т.к. в нем есть книги.")
            else:
                obj.delete()
                messages.success(request, "Жанр удален.")


        elif item_type == 'reason':

            obj = get_object_or_404(RejectionReason, id=item_id)

            # РУЧНАЯ ПРОВЕРКА ЦЕЛОСТНОСТИ:

            # Метод .booking_set.exists() проверяет, есть ли заявки, связанные с этой причиной

            if obj.booking_set.exists():

                messages.error(request,
                               f"Ошибка: Нельзя удалить причину «{obj.name}», так как она уже используется!")

            else:

                obj.delete()

                messages.success(request, "Причина отказа удалена из справочника.")

    except ProtectedError:
        messages.error(request, "Удаление невозможно: эта запись используется в других разделах.")
    except Exception as e:
        messages.error(request, f"Системная ошибка: {e}")

    # 5. Финальный редирект (теперь переменные tab и base_url точно существуют)
    return redirect(f"{base_url}?tab={tab}")


# === НОВАЯ ЛОГИКА: РЕДАКТИРОВАНИЕ СПРАВОЧНИКОВ ===
@login_required
def edit_dict_item(request, item_type, item_id):
    if request.user.role != 'librarian':
        return redirect('library:catalog')

    # Инициализация переменных по умолчанию
    obj = None
    form_class = None
    tab = 'books'
    base_url = '/catalog/' # Большинство справочников в каталоге

    # Определяем, что именно мы редактируем
    if item_type == 'author':
        obj = get_object_or_404(Author, id=item_id)
        form_class = AuthorForm
        tab = 'authors'
    elif item_type == 'genre':
        obj = get_object_or_404(Genre, id=item_id)
        form_class = GenreForm
        tab = 'genres'
    elif item_type == 'publisher':
        obj = get_object_or_404(Publisher, id=item_id)
        form_class = PublisherForm
        tab = 'publishers'
    elif item_type == 'reason': # ДОБАВИЛИ ПОДДЕРЖКУ ПРИЧИН
        obj = get_object_or_404(RejectionReason, id=item_id)
        form_class = RejectionReasonForm
        tab = 'reasons'
        base_url = '/librarian/' # Причины возвращают в журнал бронирования

    # Если тип не распознан, выходим
    if not form_class:
        return redirect('library:catalog')

    if request.method == 'POST':
        form = form_class(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Данные успешно обновлены.")
            # Возвращаемся на нужную страницу и нужную вкладку
            return redirect(f"{base_url}?tab={tab}")
    else:
        form = form_class(instance=obj)

    return render(request, 'library/edit_dict.html', {
        'form': form,
        'type': item_type,
        'base_url': base_url,
        'tab': tab
    })


@login_required
def book_create_view(request):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Книга успешно добавлена в каталог!")
            return redirect('library:catalog')
    else:
        form = BookForm()

    return render(request, 'library/book_form.html', {'form': form, 'is_edit': False})


# === РЕДАКТИРОВАНИЕ КНИГИ ===
@login_required
def book_edit_view(request, book_id):
    if request.user.role not in ['admin', 'librarian']:
        return redirect('library:catalog')

    book = get_object_or_404(Book, id=book_id)

    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, "Информация о книге обновлена!")
            return redirect('library:catalog')
    else:
        form = BookForm(instance=book)

    return render(request, 'library/book_form.html', {'form': form, 'is_edit': True, 'book': book})